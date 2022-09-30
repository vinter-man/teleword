import collections
import datetime
import logging
import random
import sys
import json
import csv
import time

import openpyxl
import sqlalchemy
import secrets
import requests
import copy
import os

import mysql.connector
from xml.etree import ElementTree
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship
from sqlalchemy.orm import sessionmaker
import sqlalchemy.exc
from sqlalchemy.exc import PendingRollbackError

from config.config import MY_SQL, APP_KEY_OXF, APP_ID_OXF, URL_OXF, ADMIN_ID_TG


########################################################################################################################
logger = logging.getLogger(__name__)
logging.basicConfig(
        level=logging.INFO,
        stream=sys.stdout,
        format='[%(asctime)s]:[%(levelname)s]:[%(filename)s]:[%(lineno)d]: %(message)s',
    )


########################################################################################################################
########################################################################################################################
# basic
engine = sqlalchemy.create_engine(MY_SQL, pool_recycle=3600, echo=True)
Base = declarative_base()    # this guy will set the trend :)


########################################################################################################################
# python classes
class Users(Base):
    """
    Table for storing data of teleword users
    """
    __tablename__ = 'users'

    user_id = sqlalchemy.Column(sqlalchemy.Integer, primary_key=True, autoincrement=True, nullable=False)
    tg_id = sqlalchemy.Column(sqlalchemy.String(20))
    nickname = sqlalchemy.Column(sqlalchemy.String(250), nullable=False)
    lang_code = sqlalchemy.Column(sqlalchemy.String(10), nullable=False)
    shock_mode = sqlalchemy.Column(sqlalchemy.Integer, default=0)
    points = sqlalchemy.Column(sqlalchemy.Integer, default=0)
    is_blacklisted = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    is_bot = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    creation_time = sqlalchemy.Column(sqlalchemy.String(20), nullable=False)
    last_use_time = sqlalchemy.Column(sqlalchemy.String(20), nullable=False)
    current_use_time = sqlalchemy.Column(sqlalchemy.String(20), nullable=False)

    exxs = relationship('UsersExamples', backref='user_examples')    # class not table name
    stats = relationship('UsersStatistics', backref='user_statistics')
    apiks = relationship('UsersApiKeys', backref='user_apikeys')


class UsersExamples(Base):
    """
    Table for storing data of user examples
    """
    __tablename__ = 'examples'

    ex_id = sqlalchemy.Column(sqlalchemy.Integer, primary_key=True, autoincrement=True, nullable=False)
    example = sqlalchemy.Column(sqlalchemy.String(400), nullable=False)
    user_id = sqlalchemy.Column(sqlalchemy.Integer, sqlalchemy.ForeignKey('users.user_id'), nullable=False)

    words = relationship('UsersExamplesWords', backref='example_words')


class UsersExamplesWords(Base):
    """
    Table for storing data of example words
    """
    __tablename__ = 'words'
    # word_id - sqlalchemy.Integer -> max 350000 users with 6000 words
    word_id = sqlalchemy.Column(sqlalchemy.Integer, primary_key=True, autoincrement=True, nullable=False)
    word = sqlalchemy.Column(sqlalchemy.String(135), nullable=False)
    description = sqlalchemy.Column(sqlalchemy.String(400), nullable=False)
    category = sqlalchemy.Column(sqlalchemy.String(20))
    rating = sqlalchemy.Column(sqlalchemy.Integer, default=0)
    example_id = sqlalchemy.Column(sqlalchemy.Integer, sqlalchemy.ForeignKey('examples.ex_id'))


class UsersStatistics(Base):
    """
    Table for storing data of user statistic
    """
    __tablename__ = 'statistics'

    day_id = sqlalchemy.Column(sqlalchemy.Integer, primary_key=True)
    day = sqlalchemy.Column(sqlalchemy.String(10), nullable=False)
    firs_try_success = sqlalchemy.Column(sqlalchemy.Integer)
    mistake = sqlalchemy.Column(sqlalchemy.Integer)
    total = sqlalchemy.Column(sqlalchemy.Integer)
    user_id = sqlalchemy.Column(sqlalchemy.Integer, sqlalchemy.ForeignKey('users.user_id'))


class UsersApiKeys(Base):
    """
    Table for storing data of user api-keys
    """
    __tablename__ = 'apikeys'

    key_id = sqlalchemy.Column(sqlalchemy.Integer, autoincrement=True, primary_key=True)
    key = sqlalchemy.Column(sqlalchemy.String(130), nullable=False)
    user_id = sqlalchemy.Column(sqlalchemy.Integer, sqlalchemy.ForeignKey('users.user_id'))


########################################################################################################################
# create
Base.metadata.create_all(engine)


########################################################################################################################
# work with
DbSession = sessionmaker(bind=engine)
session = DbSession()


########################################################################################################################
########################################################################################################################
# common.py
def is_user(telegram_id: str) -> bool:
    """
    Returns the result of checking if the user exists in the database by telegram_id
    :param telegram_id: string representation of telegram user id
    """
    result = session.query(Users).filter(Users.tg_id == telegram_id).all()
    if not result:
        return False
    return True


def add_user(tg_id: str, nickname: str, lang_code: str, shock_mode: int,
             points: int, is_blacklisted: bool, is_bot: bool, creation_time: str,
             last_use_time: str, current_use_time: str):
    """
    Adds a new user to the database

    :param tg_id: string representation of telegram user id
    :param nickname: string representation of telegram user username
    :param lang_code:  string representation of telegram user language code
    :param shock_mode: integer shock mode start value
    :param points: integer points start value
    :param is_blacklisted: bool black list status
    :param is_bot: bool is user a bot status
    :param creation_time: string representation of creation time data
    :param last_use_time: string representation of last time use data
    :param current_use_time: string representation of current use time data
    """
    session.add(Users(
        tg_id=tg_id, nickname=nickname, lang_code=lang_code,
        shock_mode=shock_mode, points=points, is_blacklisted=is_blacklisted,
        is_bot=is_bot, creation_time=creation_time,
        last_use_time=last_use_time, current_use_time=current_use_time
    ))
    session.commit()


def change_user_last_using(user_tg_id: str, flag: str = 'change'):
    """
    Checks / Increments / Resets the last time the user used the application
    :param user_tg_id: string representation of telegram user id
    :param flag: string 'change' / 'check'
    """

    user = session.query(Users).filter_by(tg_id=user_tg_id).one()
    user.current_use_time = str(datetime.date.today())

    difference = datetime.date.fromisoformat(user.current_use_time) - datetime.date.fromisoformat(user.last_use_time)
    difference = int(difference.days)

    if not difference or (difference == 1 and flag == 'check'):
        logger.info(f'| {user_tg_id} | no different user using time')
    elif difference == 1 and flag == 'change':
        logger.info(f'| {user_tg_id} | shock_mode +1 day')
        user.shock_mode += 1
        user.last_use_time = user.current_use_time
    else:
        logger.info(f'| {user_tg_id} | shock_mode is finish')
        user.shock_mode = 0
        user.last_use_time = user.current_use_time

    session.add(user)
    session.commit()


def users_bl_list() -> list:
    """
    :return: list of blacklisted users
    """
    return [i for i in session.query(Users).filter_by(is_blacklisted=True).all()]


def change_user_bl_status(user_tg_id: str, change_for: bool):
    """
    Change user black list status
    :param user_tg_id: string representation of telegram user id
    :param change_for: new bool value
    """

    user = session.query(Users).filter_by(tg_id=user_tg_id).one()

    if change_for:
        user.is_blacklisted = True
    else:
        user.is_blacklisted = False

    session.add(user)
    session.commit()
    logger.info(f'| {user_tg_id} | changed black list status to "{change_for}"')


########################################################################################################################
# adding.py
def add_example(example_text: str, user_tg_id: str) -> UsersExamples:
    """
    Adding new example to 'examples' table, return UsersExamples obj
    """

    user = get_user(tg_id=user_tg_id)
    user_id = user.user_id

    example_in = session.query(UsersExamples).filter(sqlalchemy.and_(
        UsersExamples.example == example_text,
        UsersExamples.user_id == user_id
    )).first()   # First result | None
    if example_in:      # no need to add
        logger.info(f'| {user_tg_id} | example already added return value')
        return example_in

    example_to_add = UsersExamples(
        example=example_text,
        user_id=user_id
    )
    session.add(example_to_add)
    session.commit()
    logger.info(f'| {user_tg_id} | example successfully added return value')
    return example_to_add


def add_word(word: str, description: str, category: str, rating: int, example: UsersExamples) -> UsersExamplesWords:
    """
    Adding new word to 'words' table
    """

    word_in = session.query(UsersExamplesWords).filter(sqlalchemy.and_(
        UsersExamplesWords.word == word,
        UsersExamplesWords.description == description,
        UsersExamplesWords.example_id == example.ex_id,
    )).first()   # First result | None

    if word_in:
        logger.warning(f'| {word, example.ex_id} | word already added')
        return word_in

    word_to_add = UsersExamplesWords(
        word=word,
        description=description,
        category=category,
        rating=rating,
        example_id=example.ex_id    # or "example_words=example"
    )
    session.add(word_to_add)
    session.commit()
    logger.info(f'| {word, example.ex_id} | word successfully added')
    return word_to_add


########################################################################################################################
# lessons.py
def get_user(tg_id: str) -> Users | None:
    """
    :param tg_id: string representation of telegram user id
    :return: return user Users object
    """
    return session.query(Users).filter_by(tg_id=tg_id).first()


def get_word(word_id: int) -> UsersExamplesWords | None:
    """
    :param word_id: integer representation of word id
    :return: return word UsersExamplesWords object
    """
    return session.query(UsersExamplesWords).filter_by(word_id=word_id).first()


def change_rating(word_id: int, new_rating: int):
    """
    Changes the rating of a word
    :param word_id: integer representation of word id
    :param new_rating: integer representation of new rating
    """
    word = get_word(word_id=word_id)

    if not word:
        logger.warning(f'| {word_id} | no word with user word_id')
        raise KeyError(f'{word_id}')

    old_rating = word.rating
    word.rating = new_rating
    session.add(word)
    session.commit()
    logger.info(f'changed word rating {old_rating} >>> {new_rating}')


def add_or_change_day_stat(tg_id: str, first_try: int, mistakes: int, points: int = 15):
    """
    Makes / changes the daily log in the user statistics table
    :param tg_id: string representation of telegram user id
    :param first_try: integer count of first attempts success
    :param mistakes: integer count of mistakes
    :param points: integer count of points
    """
    today = str(datetime.date.today())
    user = get_user(tg_id=tg_id)
    user_id = user.user_id

    day_stat_log = session.query(UsersStatistics).filter(sqlalchemy.and_(
        UsersStatistics.user_id == user_id,
        UsersStatistics.day == today
    )).first()

    if day_stat_log:
        day_stat_log.firs_try_success += first_try
        day_stat_log.mistake += mistakes
        day_stat_log.total += points
        logger.info(f'change day stat f_try {first_try}, mistakes {mistakes}, total {points}...')
    else:
        day_stat_log = UsersStatistics(
            day=today,
            firs_try_success=first_try,
            mistake=mistakes,
            total=points,
            user_id=user_id
        )
        logger.info(f'add day stat f_try {first_try}, mistakes {mistakes}, total {points}...')

    user.points += points

    session.add(day_stat_log)
    session.add(user)
    session.commit()
    logger.info(f'successes add_or_change_day_stat and changing total user points')


def get_words_data(user_tg_id: str) -> list:
    """
    :param user_tg_id: string representation of telegram user id
    :return: the data of the user's words by list with json-words with keys:
     'tg_id': 'word', 'description', 'example', 'category', 'rating', 'word_id', 'is_main'
    """
    sql_query = engine.execute(
        " SELECT "
        "examples.user_id, words.word, words.description, examples.example, words.category, words.rating, words.word_id"
        " FROM users"
        " LEFT JOIN examples ON examples.user_id = users.user_id"
        " LEFT JOIN words ON words.example_id = examples.ex_id"
        " WHERE users.tg_id = '{}'".format(user_tg_id)
    )
    words = [
        {
            'tg_id': i.user_id,
            'word': i.word,
            'description': i.description,
            'example': i.example,
            'category': i.category,
            'rating': i.rating,
            'word_id': i.word_id,
            'is_main': False
        }
        for i in sql_query
    ]
    logger.info(f'{user_tg_id} Successes return words data')
    return words


class MinLenError(TypeError):
    """
    It occurs when a lesson requests when the user does not have enough words
    """


def get_lesson_data(tg_id: str) -> list:
    """
    independent action
        accepts a db_worker.Users instance (further "user")
        collects the user's words
        if there are less than 15 words - raises an exception
        makes a sequence of tests depending on the rating of the word
        Adds wrong answers depending on the speech type of the word
        Returns a list filled with 15-lists of 4 words
        of dict {'tg_id', 'word', 'description', 'example', 'category', 'rating', 'word_id', 'is_main'}
    """
    logger.info(f'{tg_id} Start get_lesson_data')
    words = get_words_data(user_tg_id=tg_id)
    words_len = len(words)
    if words_len < 15:
        raise MinLenError(f"{words_len}")
    words.sort(key=lambda word: word['rating'], reverse=True)

    # the most difficult words are better learned 1/3(5):
    # 1 - 3 when repeated at the beginning (as a work on mistakes)
    # 14 - 15 and as the last tasks (like a boss in a video game,
    # so that the player has fun after passing)
    difficult_words = words[:5]
    random.shuffle(difficult_words)
    remaining_words = words[5:]
    random.shuffle(remaining_words)
    easy_words = remaining_words[:10]
    words_for_lesson = collections.deque(
        iterable=difficult_words[:3] + easy_words + difficult_words[3:],
        maxlen=15
    )
    data_for_lesson = []
    # data for the test 15 tests (1 correct and 3 wrong answers)
    for i in range(len(words_for_lesson)):
        random.shuffle(words)  # i mix each time, to improve accident
        test_words = []  # 4 words, in the first place is always correct

        right_word = words_for_lesson.popleft()
        test_words.append(right_word)

        for word in words:
            if len(test_words) == 4:  # 4 words
                break
            if word not in test_words and word["category"] == right_word["category"]:
                # the best learning effect is when the words are not just random,
                # but belong to the same part of speech ...
                wrong_word = word
                test_words.append(wrong_word)

        if len(test_words) != 4:  # ... but the user does not always have enough words ...
            for word in words:
                if len(test_words) == 4:
                    break
                if word not in test_words:
                    # ... therefore, fill in the missing ones in order,
                    # random.shuffle at the beginning will prevent repetitions
                    wrong_word = word
                    test_words.append(wrong_word)

        data_for_lesson.append(test_words)

    # [[{w}, {w}, {w}, {w}], ...]

    # It may be that the user has added the same word with different examples,
    # in this case there will be several correct answers.
    # This approach is quite good because the word will be remembered in several contexts
    ready_tasks = []
    for task in data_for_lesson:
        task = copy.deepcopy(task)     # Each test should not affect the state of all words in general
        main_correct_word = task[0]
        main_correct_word["is_main"] = True
        correct_pattern = main_correct_word["word"].lower().strip()
        for w in task:
            w_pattern = w["word"].lower().strip()
            if w_pattern == correct_pattern:
                w["is_correct"] = True
            else:
                w["is_correct"] = False

        current_lesson_data = {'a': None, 'b': None, 'c': None, 'd': None}
        random.shuffle(task)  # answer options mixed
        for n, k in enumerate(current_lesson_data.keys()):
            current_lesson_data[k] = task[n]

        ready_tasks.append(current_lesson_data)

    # [{aw, bw, cw, dw}, {}, {}, {}, ...]
    logger.info(f'{tg_id} Return data for lesson')
    return ready_tasks


########################################################################################################################
# statistic.py
def word_count(user_tg_id: str) -> int:
    """
    :param user_tg_id: string representation of telegram user id
    :return: integer value user's word count
    """
    return (session.query(Users, sqlalchemy.func.count(UsersExamples.ex_id))
                .outerjoin(UsersExamples)
                .group_by(Users)
                .outerjoin(UsersExamplesWords)
                .group_by(Users)
                .filter(Users.tg_id == user_tg_id)
            ).one()[1]


def get_user_stat(user_tg_id: str, limit: int = 7) -> list:
    """
    :param user_tg_id: string representation of telegram user id
    :param limit: integer representation of maximum count of statistic logs
    :return: list of UsersStatistics objects
    """
    return session.query(UsersStatistics).filter_by(user_id=get_user(user_tg_id).user_id).order_by(
                                                                            UsersStatistics.day_id.desc()).limit(limit)


def build_total_mistakes_firsttry_data_for_graph(user_sql_logs: list, future_length: int = 7) -> dict:
    """
    :param user_sql_logs: list of UsersStatistics objects
    :param future_length: length of data to graph building
    :return: dict of 3 lists: 'total', 'mistakes', 'first_try'
    """
    result_dict = {
        "total": [0] * future_length,
        "mistakes": [0] * future_length,
        "first_try": [0] * future_length
    }

    today = datetime.date.today()
    for stat in user_sql_logs:
        stat_date = datetime.date.fromisoformat(stat.day)
        difference = today - stat_date
        difference = int(difference.days)
        if difference >= future_length:     # old record
            continue
        else:
            place = (future_length - difference) - 1
            result_dict.get("total")[place] = stat.total
            result_dict.get("mistakes")[place] = stat.mistake
            result_dict.get("first_try")[place] = stat.firs_try_success

    return result_dict


########################################################################################################################
# checking.py
def create_file_with_user_words(user_tg_id: str, file_path: str, file_type: str,
                                sql_filter_key: str, sql_sort_key: str) -> str:
    """
    Creates a file in the specified format with user data
    :param user_tg_id: string representation of telegram user id
    :param file_path: string place where file will be safe
    :param file_type: 'xlsx' / 'json' / 'xml' / 'csv'
    :param sql_filter_key: 'most important words' / 'default'
    :param sql_sort_key: 'by importance' / 'in alphabetical order' / 'default'
    :return: path with name of created file
    """
    # sql_filter_key:
    if sql_filter_key == 'most important words':
        sql_main = " SELECT " \
                   " words.word_id, words.word, words.description, examples.ex_id, examples.example" \
                   " FROM users" \
                   " LEFT JOIN examples ON examples.user_id = users.user_id" \
                   " LEFT JOIN words ON words.example_id = examples.ex_id" \
                   " WHERE users.tg_id = '{}' AND words.rating > (SELECT AVG(rating) FROM words)".format(user_tg_id)
    # elif sql_filter_key == '...':    # * space for expansion
    #     pass
    else:
        sql_main = " SELECT " \
                   " words.word_id, words.word, words.description, examples.ex_id, examples.example" \
                   " FROM users" \
                   " LEFT JOIN examples ON examples.user_id = users.user_id" \
                   " LEFT JOIN words ON words.example_id = examples.ex_id" \
                   " WHERE users.tg_id = '{}'".format(user_tg_id)

    # sql_sort_key:
    if sql_sort_key == 'by importance':
        sql_query = engine.execute(sql_main + ' ORDER BY words.rating ASC')
    elif sql_sort_key == 'in alphabetical order':
        sql_query = engine.execute(sql_main + ' ORDER BY words.word ASC')
    else:
        sql_query = engine.execute(sql_main + ' ORDER BY words.word_id ASC')

    # file_type:
    file_path += '/'
    if not os.path.isdir(file_path):
        os.mkdir(file_path)
    file_name = str(file_path + 'words' + user_tg_id + '.' + file_type)
    file = open(file_name, 'w', encoding='utf-8')
    # xlsx
    if file_name.endswith('xlsx'):
        workbook = openpyxl.Workbook()
        new_sheet = workbook.active
        sheet = workbook.create_sheet(f"words")
        sheet.column_dimensions['A'].width = 8.43
        sheet.column_dimensions['B'].width = 11
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 11
        sheet.column_dimensions['F'].width = 50
        sheet.column_dimensions['G'].width = 150
        workbook.remove(new_sheet)
        # write data
        for number, i in enumerate(sql_query, start=2):
            if number == 2:
                sheet[f'B{number}'] = "word id"
                sheet[f'C{number}'] = "word"
                sheet[f'D{number}'] = "description"
                sheet[f'E{number}'] = "ex id"
                sheet[f'F{number}'] = "example"
            # body
            sheet[f'B{number + 1}'] = i.word_id
            sheet[f'C{number + 1}'] = i.word
            sheet[f'D{number + 1}'] = i.description
            sheet[f'E{number + 1}'] = i.ex_id
            sheet[f'F{number + 1}'] = i.example
        # for design and fun
        sheet[f'V{number + 150}'] = 'you found the easter egg:)'
        # design
        my_font = Font(
            name='Segoe UI',
            size=14,
            color='B3BAC0'
        )
        my_font_blue = Font(
            name='Segoe UI',
            size=14,
            color='62ACBE'
        )
        my_fill = PatternFill(
            "solid",
            fgColor="1E252B")
        my_border = Border(
            left=Side(border_style='thin', color='0D0D0D'),
            right=Side(border_style='thin', color='0D0D0D'),
            top=Side(border_style='thin', color='0D0D0D'),
            bottom=Side(border_style='thin', color='0D0D0D')
        )
        my_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )
        for row in sheet.iter_rows():
            for cell in row:
                cell.fill = my_fill
                cell.font = my_font
                cell.alignment = my_alignment
                # table-only borders:
                if (
                    cell.coordinate[0] in 'B C D E F'.split()
                    and cell.coordinate[1].isdigit()
                ):
                    if 2 <= int(cell.coordinate[1:]) <= number + 1:
                        cell.border = my_border
                # blue color for indices and header:
                if (
                    (cell.coordinate[0] in 'B C D E F'.split()
                     and cell.coordinate[1:] == '2')
                    or
                    (cell.coordinate[0] in 'B E'.split())
                ):
                    cell.font = my_font_blue
        workbook.save(file_name)
    # json
    elif file_name.endswith('json'):
        json_data = {}
        for i in sql_query:
            json_data[i.word_id] = {
                "word": i.word,
                "description": i.description,
                "example_id": str(i.ex_id),
                "example": i.example
            }
        json.dump(json_data, file, indent=4, ensure_ascii=False)
    # xml
    elif file_name.endswith('xml'):
        root = ElementTree.Element('words')
        for i in sql_query:
            w = ElementTree.SubElement(root, 'w')
            word_id = ElementTree.SubElement(w, 'word_id')
            word_id.text = str(i.word_id)
            word = ElementTree.SubElement(w, 'word')
            word.text = i.word
            description = ElementTree.SubElement(w, 'description')
            description.text = i.description
            example_id = ElementTree.SubElement(w, 'example_id')
            example_id.text = str(i.ex_id)
            example = ElementTree.SubElement(w, 'example')
            example.text = i.example
        tree = ElementTree.ElementTree(root)
        tree.write(file_name)
    # сsv
    elif file_name.endswith('csv'):
        writer = csv.writer(file, quoting=csv.QUOTE_ALL)
        word_data = [(i.word_id, i.word, i.description, i.ex_id, i.example) for i in sql_query]
        writer.writerows(word_data)
    # unknown
    else:
        file.close()
        raise NameError(f'fyle type "{file_name}" is not defined')
    file.close()

    logger.info(f'{user_tg_id} Successes return new temporary user file path')
    return file_name


########################################################################################################################
# updating.py
def get_user_example(user: Users, example_id: int = None, example: str = None) -> UsersExamples | None:
    """
    :param user: user Users object
    :param example_id: integer example id
    :param example: string user example
    :return: first example user by the available parameters by UsersExamles object
    """
    user_examples_id = list(map(lambda ex: ex.ex_id, user.exxs))
    if example:
        return session.query(UsersExamples).filter(sqlalchemy.and_(
            UsersExamples.example == example, UsersExamples.ex_id.in_(user_examples_id))).first()
    else:
        return session.query(UsersExamples).filter(sqlalchemy.and_(
            UsersExamples.ex_id == example_id, UsersExamples.ex_id.in_(user_examples_id))).first()


def get_example(example_id: int) -> UsersExamples | None:
    """
    :param example_id: integer example id
    :return: first example by the available parameters by UsersExamles object
    """
    return session.query(UsersExamples).filter_by(ex_id=example_id).first()


def get_user_word(user: Users, word_id: int = None,
                  word: str = None, description: str = None) -> UsersExamplesWords | None:
    """
    :param user: user Users object
    :param word_id: integer word id
    :param word: string user word
    :param description: string user description
    :return: first user word by the available parameters in UsersExamles object form
    """
    user_examples_id = list(map(lambda example: example.ex_id, user.exxs))
    if word:
        return session.query(UsersExamplesWords).filter(sqlalchemy.and_(
            UsersExamplesWords.word == word, UsersExamplesWords.example_id.in_(user_examples_id))).first()
    elif description:
        return session.query(UsersExamplesWords).filter(sqlalchemy.and_(
            UsersExamplesWords.description == description, UsersExamplesWords.example_id.in_(user_examples_id))).first()
    else:
        return session.query(UsersExamplesWords).filter(sqlalchemy.and_(
            UsersExamplesWords.word_id == word_id, UsersExamplesWords.example_id.in_(user_examples_id))).first()


def get_word_category(word: str, default='-', url=URL_OXF) -> str:
    """
    :param word: string word
    :param default: return it if no find result
    :param url: api oxford url
    :return: string word category by the available parameters
    """
    url += word.lower()
    headers = {
        'app_id': APP_ID_OXF,
        'app_key': APP_KEY_OXF
    }
    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        logger.warning(f'{word} Requests error '
                       f'{r.status_code, headers, url} \n {r.text} \n ')
        return default

    word_data = r.json()
    if "error" in word_data.keys():
        logger.warning(f'{word} No entry found that matches the provided data'
                       f'{r.status_code, headers, url} \n {r.text} \n ')
        return default

    # Noun | Verb ... (Part of speech)
    return word_data["results"][0]["lexicalEntries"][0]["lexicalCategory"]["text"]


def update_data(data_type: str, data_id: int, new_data: str):
    """
    Update data to new values
    :param data_type: 'word' / 'description' / 'example'
    :param data_id: integer id of data
    :param new_data: string new data
    """
    if data_type == 'word':
        word = get_word(word_id=data_id)
        word.word = new_data
        word.category = get_word_category(word=new_data)
        session.add(word)
    elif data_type == 'description':
        word = get_word(word_id=data_id)
        word.description = new_data
        session.add(word)
    elif data_type == 'example':
        example = get_example(example_id=data_id)
        example.example = new_data
        session.add(example)
    else:
        raise ValueError(f'TypeError can only work with "word", "description", "example" (not "{data_type}")')
    session.commit()


def delete_data(data_type: str, data_id: int):
    """
    Delete data from id
    :param data_type: 'word' / 'description' / 'example'
    :param data_id: integer data id
    """
    if data_type in ('word', 'description'):
        word = get_word(word_id=data_id)
        word_example = get_example(example_id=word.example_id)
        session.delete(word)
        if len(word_example.words) == 0:
            session.delete(word_example)
    elif data_type == 'example':
        example = get_example(example_id=data_id)
        words = example.words
        for word in words:
            session.delete(word)
        session.delete(example)
    else:
        raise ValueError(f'TypeError can only work with "word", "description", "example" (not "{data_type}")')
    session.commit()


########################################################################################################################
# api.py
def generate_api_keys(user: Users):
    """
    Create new api-key to user
    :param user: user Users object
    """
    while True:
        new_key = secrets.token_urlsafe(64)
        key = session.query(UsersApiKeys).filter(sqlalchemy.and_(
            UsersApiKeys.key == new_key
        )).first()
        if not key:
            break
    session.add(UsersApiKeys(
        key=new_key,
        user_id=user.user_id
    ))
    session.commit()


def is_api_keys(user: Users) -> bool:
    """
    :param user: user Users object
    :return: boolean result of checking if there is such a key in the table
    """
    api_key = session.query(UsersApiKeys).filter(sqlalchemy.and_(
        UsersApiKeys.user_id == user.user_id
    )).first()
    if api_key:
        return True
    else:
        return False


def get_user_api_key(user: Users) -> str:
    """
    :param user: user Users object
    :return: string user api key
    """
    api_key: UsersApiKeys = session.query(UsersApiKeys).filter(sqlalchemy.and_(
        UsersApiKeys.user_id == user.user_id
    )).first()
    return api_key.key


def get_user_by_api_key(token: str) -> Users:
    """
    :param token: string user api key
    :return: user Users object
    """
    api_obj: UsersApiKeys = session.query(UsersApiKeys).filter(sqlalchemy.and_(
        UsersApiKeys.key == token
    )).first()
    user: Users = session.query(Users).filter(sqlalchemy.and_(
        Users.user_id == api_obj.user_id
    )).first()
    return user


########################################################################################################################
# teleword.py
def is_connection_alive() -> bool:
    """
    :return: boolean result of checking if connection with sql is alive
    """
    try:
        session.query(Users).filter_by(tg_id=ADMIN_ID_TG).first()
    except mysql.connector.errors.OperationalError as e:
        logger.error(f'Connection is dead {e}')
        return False
    return True


def pending_rollback(username: str):
    """
    Try to connect with db, if can`t - try to rollback
    :param username: string user name for logs
    """
    start = time.perf_counter()
    while True:
        if time.perf_counter() - start > 4.5:
            logger.critical(f'[{username}]: Connection with db died. Can`t reconnect')
            break
        try:
            session.commit()
            break
        except PendingRollbackError as e:
            logger.error(f'[{username}]: Connection with db died {e}. Try to reconnect...')
            session.rollback()
